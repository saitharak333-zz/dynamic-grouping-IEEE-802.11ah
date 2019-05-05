import os
from openpyxl import load_workbook
import matplotlib.pyplot

sta = 10
cou = 1
stalis = []
thr = []
count = 1
while sta <= 50:
    stalis.append(sta)
    cmd = 'python3 stationmain.py'
    cmd = cmd + ' ' + str(count) + ' ' + str(sta)
    os.system(cmd)
    sta = sta + 5
    count += 1
else:
    wb = load_workbook('station.xlsx')
    sheet = wb.active
    for i in range(len(stalis)):
        thr.append(float(sheet.cell(row = i+1, column=2).value))
    else:
        plt.title("Throughput vs No of Stations")
        plt.xlabel("No of stations")
        plt.ylabel("Throughput")
        plt.plot(stalis, thr)
        plt.show()
        print('finished')
