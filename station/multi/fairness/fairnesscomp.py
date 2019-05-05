import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt

fair1 = []
fair2 = []
x = [i+1 for i in range(100)]
wb = load_workbook('fairness_100.xlsx')
sheet = wb.active
for i in range(100):
    fair1.append(float(sheet.cell(row = i+1, column=1).value))
    fair2.append(float(sheet.cell(row = i+1, column=2).value))
else:
    plt.title("Fairness Comparision")
    plt.plot(x, fair1, label = 'Dynamic')
    plt.plot(x, fair2, label = 'DCF')
    plt.legend()
    plt.ylabel("Fairness")
    plt.xlabel("Beacon Interval")
    plt.show()
    print('finished')
