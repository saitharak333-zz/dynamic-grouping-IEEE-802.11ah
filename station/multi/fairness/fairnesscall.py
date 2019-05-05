import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt

sta = 10
stalis = []
fair = []
count = 1
while sta <= 150:
    stalis.append(sta)
    cmd = 'python3 fairnessmain.py'
    cmd = cmd + ' ' + str(count) + ' ' + str(sta)
    os.system(cmd)
    sta = sta + 5
    count += 1
# else:
    # wb = load_workbook('beacon.xlsx')
    # sheet = wb.active
    # for i in range(len(bealis)):
    #     thr.append(float(sheet.cell(row = i+1, column=2).value))
    # else:
    #     plt.title("Throughput vs Beacon")
    #     plt.xlabel("No of stations")
    #     plt.ylabel("Beacon Interval")
    #     plt.plot(bealis, thr)
    #     plt.show()
    #     print('finished')
