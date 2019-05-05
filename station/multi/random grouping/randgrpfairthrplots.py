import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt

sta = 10
thrrand = []
thrdy = []
fairrand = []
fairdy = []
bea = [i+1 for i in range(100)]
wb = load_workbook('randfair.xlsx')
sheet = wb.active
for i in range(100):
    thrrand.append(float(sheet.cell(row = i+1, column=2).value))
    thrdy.append(float(sheet.cell(row = i+1, column=4).value))
    fairrand.append(float(sheet.cell(row = i+1, column=1).value))
    fairdy.append(float(sheet.cell(row = i+1, column=3).value))
else:
    plt.title("Fairness Comparision")
    plt.xlabel("Beacon Intervals")
    plt.ylabel("Fairness")
    plt.figure(1)
    plt.ylim(0,1)
    plt.plot(bea, fairrand, label = 'Uniform')
    plt.plot(bea, fairdy, label = 'Dynamic')
    plt.legend()
    plt.show()


    plt.figure(2)
    plt.title("Throughput Comparision")
    plt.xlabel("Beacon Intervals")
    plt.ylabel("Throughput")
    plt.ylim(0,1)
    plt.plot(bea, thrrand, label = 'Uniform')
    plt.plot(bea, thrdy, label = 'Dynamic')
    plt.legend()
    plt.show()

    print('finished')
