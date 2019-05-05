import random
import time
import threading
import statistics
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Reading the excel file
# import pandas as pd
# df = pd.read_excel('dataset.xlsx')
# df_norm = (df - df.mean())/(df.max() - df.min())
# # Mean values of dataset
# stationMean = 265
# slotsMean = 25.642857
# BeaconIntervalMean = 121192.087912
# ThroughputMean = 0.278645
# # max - min of Dataset
# stationMaxmin = 470
# slotsMaxmin = 49
# BeaconIntervalMaxmin = 163840
# ThroughputMaxmin = 0.264033
# # Dataset for no of slots
# slots_x = df_norm[['No of Stations', 'Beacon Interval', 'Throughput']]
# slots_y = df_norm[['No of Slots']]
# # Dataset for Throughput
# th_x = df_norm[['No of Stations', 'Beacon Interval', 'No of Slots']]
# th_y = df_norm[['Throughput']]
# # Regression models
# from sklearn.tree import DecisionTreeRegressor
# reg_slots = DecisionTreeRegressor(max_depth = 7)
# reg_th = DecisionTreeRegressor(max_depth = 20)
# reg_slots.fit(slots_x, slots_y)
# reg_th.fit(th_x, th_y)

# Threading function call for each station
def mac(lock, station_id, slottime):
    begin = time.time()  # time when thread got started
    global pktrec       #  Packets recieved for each station
    global channel      #  Channel is set as global
    CWmin = 15          #  Min window size
    CWmax = 1024        #  Max window size
    while True:
        if time.time() - begin > slottime:
            break
        count = 0         # a variable to keep track of retrylimit and use in selecting backoff value
        while True:
            if time.time() - begin > slottime:
                break
            # Selecting a random backoff counter value
            count += 1
            backOff = int(random.randint(1,min((pow(2,count)-1) * CWmin, CWmax)))
            while backOff > 0:
                if time.time() - begin > slottime:
                    break
                # Waste is for sigma is taken as 50us
                waste = [0 for i in range(1000)]

                if channel == 0:
                    backOff = backOff - 1
            else:
                while channel == 1:
                    if time.time() - begin > slottime:
                        break
                    waste = [0 for i in range(1000)]
                if time.time() - begin > slottime:
                    break
                lock.acquire()
                if channel == 0:
                    channel = 1
                else:
                    while channel:
                        waste = [0 for i in range(1000)]
                    else:
                        channel = 1
                lock.release()
                if time.time() - begin > slottime:
                    channel = 0
                    break
                time.sleep(250 * (10**(-6)))    #  Time for sending the packet
                if time.time() - begin > slottime:
                    channel = 0
                    break
                pktrec[station_id] += 1
                channel = 0

# Number of Stations
N = int(input("No of Nodes connected to the server: "))
# Data Rate
# datarate = float(input("Enter the data rate in Mbps: "))
# data rate is taken as 0.65 Mbps
# Packet Size
size = int(input("Enter the packet size in bytes: "))
# Packet Arrival Rate
# Saturates traffic scenario is considered
# pktrate = int(input("Enter packet arrival rate: "))
# Beacon Interval
beaconinterval = int(input("Enter Beacon Interval in us: "))
beaconinterval = float(beaconinterval * (10 ** (-6)))
# Channel
channel = 0
# 0 means channel idle
# 1 means channel busy

# Initialisation
# No of packets recieved from each station
pktrec = [0 for i in range(N)]
lock = threading.Lock()     #  Used for locking the channel
notimes = 1
# wb = load_workbook('randfair.xlsx')
# sheet = wb.active

while notimes != 21:
    # No of groups
    nogrp = int(random.randint(1,N))
    # Group duration
    grpdur = []
    for i in range(nogrp):
        grpdur.append(float(beaconinterval/nogrp))
    # Grouping of Stations into groups
    grpsta = [[] for i in range(nogrp)]
    for i in range(N):
        grpsta[i % nogrp].append(i)
    # No of raw slots in each group
    slotspergrp = []
    for i in range(nogrp):
        rawslot = int(random.randint(1,len(grpsta[i])))
        slotspergrp.append(rawslot)

    # Only for the first time
    # Itereating through all the groups and setting stations per slot and slot time

    # running the threads with the initial parameters
    for i in range(nogrp):
    # Stations present in each slot
        slotsta = [[] for j in range(slotspergrp[i])]
        for j in range(len(grpsta[i])):
            slotsta[j % slotspergrp[i]].append(grpsta[i][j])
    # Time duration of each raw slot
        slottime = [float(grpdur[i]/slotspergrp[i]) for j in range(slotspergrp[i])]
    # Iterating in each RAW slot
        for j in range(slotspergrp[i]):
        # running stations in a slot no. of threads parallelly
            for k in slotsta[j]:
                stationthreads = threading.Thread(target = mac, name = str(i), args=(lock, k, slottime[j],))
                stationthreads.start()
            else:
            # to resume execution of main function after last thread is completed
                stationthreads.join()
    else:
        thr = (sum(pktrec) * 256 * 8)/(notimes * 0.65 * 1024 * 1024 * beaconinterval)
        fairness =  1 - (statistics.stdev(pktrec)/statistics.mean(pktrec))
        # sheet.cell(row = notimes, column=1).value = fairness
        # sheet.cell(row = notimes, column=2).value = thr
        print(notimes)
        print("throughput " + str(thr))
        print("fairness " + str(fairness))
        print(pktrec)
        print(' ')
        notimes += 1
# wb.save('randfair.xlsx')
