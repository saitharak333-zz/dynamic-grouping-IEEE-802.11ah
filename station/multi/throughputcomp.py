import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt

th1 = []
th2 = []
x = [i+1 for i in range(100)]
wb = load_workbook('fairness_100.xlsx')
sheet = wb.active
for i in range(100):
    th1.append(float(sheet.cell(row = i+1, column=3).value))
    th2.append(float(sheet.cell(row = i+1, column=4).value))
else:
    plt.title("Throughput Comparision")
    plt.ylabel("Normalised Throughput")
    plt.xlabel("Beacon Interval")
    plt.plot(x, th1, label = 'Dynamic')
    plt.plot(x, th2, label = 'DCF')
    plt.legend()
    plt.show()
    print('finished')
