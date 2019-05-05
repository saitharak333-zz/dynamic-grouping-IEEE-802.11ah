import random
import time
import threading
import statistics
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Reading the excel file
import pandas as pd
df = pd.read_excel('dataset.xlsx')
df_norm = (df - df.mean())/(df.max() - df.min())
# Mean values of dataset
stationMean = 265
slotsMean = 25.642857
BeaconIntervalMean = 121192.087912
ThroughputMean = 0.278645
# max - min of Dataset
stationMaxmin = 470
slotsMaxmin = 49
BeaconIntervalMaxmin = 163840
ThroughputMaxmin = 0.264033
# Dataset for no of slots
slots_x = df_norm[['No of Stations', 'Beacon Interval', 'Throughput']]
slots_y = df_norm[['No of Slots']]
# Dataset for Throughput
th_x = df_norm[['No of Stations', 'Beacon Interval', 'No of Slots']]
th_y = df_norm[['Throughput']]
# Regression models
from sklearn.tree import DecisionTreeRegressor
reg_slots = DecisionTreeRegressor(max_depth = 7)
reg_th = DecisionTreeRegressor(max_depth = 20)
reg_slots.fit(slots_x, slots_y)
reg_th.fit(th_x, th_y)

# Threading function call for each station
def mac(lock, station_id, slottime):
    begin = time.time()  # time when thread got started
    global pktrec       #  Packets recieved for each station
    global channel      #  Channel is set as global
    CWmin = 15          #  Min window size
    CWmax = 1024        #  Max window size
    while True:
        if time.time() - begin > slottime:
            break
        count = 0         # a variable to keep track of retrylimit and use in selecting backoff value
        while True:
            if time.time() - begin > slottime:
                break
            # Selecting a random backoff counter value
            count += 1
            backOff = int(random.randint(0,min((pow(2,count)-1) * CWmin, CWmax)))
            while backOff > 0:
                if time.time() - begin > slottime:
                    break
                # Waste is for sigma is taken as 50us
                waste = [0 for i in range(1000)]

                if channel == 0:
                    backOff = backOff - 1
            else:
                while channel == 1:
                    if time.time() - begin > slottime:
                        break
                    waste = [0 for i in range(1000)]
                if time.time() - begin > slottime:
                    break
                lock.acquire()
                if channel == 0:
                    channel = 1
                else:
                    while channel:
                        waste = [0 for i in range(1000)]
                    else:
                        channel = 1
                lock.release()
                if time.time() - begin > slottime:
                    channel = 0
                    break
                time.sleep(250 * (10**(-6)))    #  Time for sending the packet
                if time.time() - begin > slottime:
                    channel = 0
                    break
                pktrec[station_id] += 1
                channel = 0

# Number of Stations
N = int(input("No of Nodes connected to the server: "))
# Data Rate
# datarate = float(input("Enter the data rate in Mbps: "))
# data rate is taken as 0.65 Mbps
# Packet Size
size = int(input("Enter the packet size in bytes: "))
# Packet Arrival Rate
# Saturates traffic scenario is considered
# pktrate = int(input("Enter packet arrival rate: "))
# Beacon Interval
beaconinterval = int(input("Enter Beacon Interval in us: "))
beaconinterval = float(beaconinterval * (10 ** (-6)))
# Channel
channel = 0
# 0 means channel idle
# 1 means channel busy

# Initialisation
# No of groups
print("For Initialisation")
nogrp = int(input("No of groups to be present: "))
# Group duration
grpdur = []
for i in range(nogrp):
    grpdur.append(float(beaconinterval/nogrp))
# No of packets recieved from each station
pktrec = [0 for i in range(N)]
# Grouping of Stations into groups
grpsta = [[] for i in range(nogrp)]
for i in range(N):
    grpsta[i % nogrp].append(i)
# No of raw slots in each group
rawslot = int(input("Enter intial no of raw slots: "))
# No of RAW Slots for each group
slotspergrp = [rawslot for i in range(nogrp)]

# Only for the first time
# Itereating through all the groups and setting stations per slot and slot time
lock = threading.Lock()     #  Used for locking the channel

# running the threads with the initial parameters
for i in range(nogrp):
    # Stations present in each slot
    slotsta = [[] for j in range(slotspergrp[i])]
    for j in range(len(grpsta[i])):
        slotsta[j % slotspergrp[i]].append(grpsta[i][j])
    # Time duration of each raw slot
    slottime = [float(grpdur[i]/slotspergrp[i]) for j in range(slotspergrp[i])]
    # Iterating in each RAW slot
    for j in range(slotspergrp[i]):
        # running stations in a slot no. of threads parallelly
        for k in slotsta[j]:
            stationthreads = threading.Thread(target = mac, name = str(i), args=(lock, k, slottime[j],))
            stationthreads.start()
        else:
            # to resume execution of main function after last thread is completed
            stationthreads.join()

print(pktrec)
notimes = 0
# wb = load_workbook('fairness_100.xlsx')
# sheet = wb.active
itr = 20;
avg = 0;
fairness = [0 for i in range(20)]
# carrying out hierarchical clustering
while itr > 0:
    # initialisation for first stage of hierarchical clustering
    grpsta = [[i] for i in range(N)]
    mpktrec = [x+1 if x == 0 else x for x in pktrec]
    invpacketrec = [float(1/x) for x in mpktrec]
    criterion = [float(x/sum(invpacketrec)) for x in invpacketrec]
    grpdur = [(beaconinterval * x) for x in criterion]
    slotspergrp = [1 for i in range(len(grpsta))]
    throughoverall = 0
    #predicting the overall throughput of the first stage
    for i in range(len(grpsta)):
        # normalising the values so that they can be fed to the regressor
        station_norm = ((len(grpsta[i]) - stationMean)/stationMaxmin)
        beaconinterval_norm = ((grpdur[i] - BeaconIntervalMean)/BeaconIntervalMaxmin)
        slotspergrp[i] = 1
        slots_norm = ((slotspergrp[i] - slotsMean)/slotsMaxmin)
        # creating a dataframe inorder to predict the throughput from no.of stations,beacon interval and no.of slots
        data = {'No of Stations':[station_norm],'Beacon Interval':[beaconinterval_norm],'No of Slots':[slots_norm]}
        ip_data = pd.DataFrame(data)
        th_norm = reg_th.predict(ip_data)
        th = (th_norm * ThroughputMaxmin) + ThroughputMean
        throughoverall += th * grpdur[i]
    else:
        finalthroughput = throughoverall/beaconinterval
        finalgrpsta = [x for x in grpsta]
        finalgrpdur = [x for x in grpdur]
        finalslotspergrp = [x for x in slotspergrp]
        # grouping till we group all the stations into one group
    while len(grpsta) != 1:
        dic = {}
        # creating a dictionary with same no.of packets received into one list
        for i in range(len(grpsta)):
            try:
                if len(dic[mpktrec[i]]):
                    dic[mpktrec[i]].append(i)
            except:
                dic[mpktrec[i]] = [i]
        else:
            # sorting the array in order of increasing packets received
            mpktrec_sorted = sorted(mpktrec)
            # finding minimum difference between the packets received and the corresponding indices
            min_diff = mpktrec_sorted[-1] - mpktrec_sorted[0]
            index = 0
            for i in range(len(mpktrec_sorted)-1):
                if min_diff > mpktrec_sorted[i + 1] - mpktrec_sorted[i]:
                    min_diff = mpktrec_sorted[i + 1] - mpktrec_sorted[i]
                    index = i
                if min_diff == 0:
                    break
        # Finding the first and second list of groups to be grouped further
        if len(dic[mpktrec_sorted[index]]) != 1:
            first = dic[mpktrec_sorted[index]][0]
            second = dic[mpktrec_sorted[index]][1]
        else:
            first = dic[mpktrec_sorted[index]][0]
            second = dic[mpktrec_sorted[index + 1]][0]
            # print(first, second)
        fir_grp = grpsta[first]
        sec_grp = grpsta[second]
        fir_grp_mpktrec = mpktrec[first]
        sec_grp_mpktrec = mpktrec[second]
        # removing the two groups from the grpsta array
        grpsta.remove(fir_grp)
        grpsta.remove(sec_grp)
        mpktrec.remove(fir_grp_mpktrec)
        mpktrec.remove(sec_grp_mpktrec)
        finalgrp = fir_grp + sec_grp
        # finding the final packet recieved by taking a weighted average
        finalgrp_mpktrec = ((len(fir_grp) * fir_grp_mpktrec) + (len(sec_grp) * sec_grp_mpktrec))/(len(fir_grp) + len(sec_grp))
        # appending the new grp formed
        grpsta.append(finalgrp)
        mpktrec.append(finalgrp_mpktrec)

        invpacketrec = [float(1/x) for x in mpktrec]
        criterion = [float(x/sum(invpacketrec)) for x in invpacketrec]
        grpdur = [(beaconinterval * x) for x in criterion]
        slotspergrp = [1 for i in range(len(grpsta))]
        throughoverall = 0

        for i in range(len(grpsta)):
            # normalising the values so that they can be fed to the regressor
            station_norm = ((len(grpsta[i]) - stationMean)/stationMaxmin)
            beaconinterval_norm = ((grpdur[i] - BeaconIntervalMean)/BeaconIntervalMaxmin)
            # creating a dataframe inorder to predict the no.of slots from stations,beacon interval and throughput
            data = {'No of Stations':[station_norm],'Beacon Interval':[beaconinterval_norm],'Throughput':[((0.5 - ThroughputMean)/ThroughputMaxmin)]}
            slots_norm = reg_slots.predict(ip_data)
            ip_data = pd.DataFrame(data)
            slotspergrp[i] = min(int((slots_norm * slotsMaxmin) + slotsMean), len(grpsta[i]))
            if slotspergrp[i] == 0:
                slotspergrp[i] = 1
            slots_norm = ((slotspergrp[i] - slotsMean)/slotsMaxmin)
            # creating a dataframe inorder to predict the throughput from stations,beacon interval and no.of slots
            data = {'No of Stations':[station_norm],'Beacon Interval':[beaconinterval_norm],'No of Slots':[slots_norm]}
            ip_data = pd.DataFrame(data)
            th_norm = reg_th.predict(ip_data)
            th = (th_norm * ThroughputMaxmin) + ThroughputMean
            throughoverall += th * grpdur[i]
        else:
            if finalthroughput < throughoverall/beaconinterval:
                finalthroughput = throughoverall/beaconinterval
                finalgrpdur = [x for x in grpdur]
                finalgrpsta = [x for x in grpsta]
                finalslotspergrp = [x for x in slotspergrp]
    else:
        grpdur = [beaconinterval]
        slotspergrp = [1]
        throughoverall = 0
        station_norm = ((len(grpsta[0]) - stationMean)/stationMaxmin)
        beaconinterval_norm = ((grpdur[0] - BeaconIntervalMean)/BeaconIntervalMaxmin)
        # creating a dataframe inorder to predict the no.of slots from stations,beacon interval and throughput
        data = {'No of Stations':[station_norm],'Beacon Interval':[beaconinterval_norm],'Throughput':[((0.5 - ThroughputMean)/ThroughputMaxmin)]}
        slots_norm = reg_slots.predict(ip_data)
        ip_data = pd.DataFrame(data)
        slotspergrp[0] = min(int((slots_norm * slotsMaxmin) + slotsMean), len(grpsta[i]))
        if slotspergrp[0] == 0:
            slotspergrp[0] = 1
        slots_norm = ((slotspergrp[0] - slotsMean)/slotsMaxmin)
        # creating a dataframe inorder to predict the throughput from stations,beacon interval and no.of slots
        data = {'No of Stations':[station_norm],'Beacon Interval':[beaconinterval_norm],'No of Slots':[slots_norm]}
        ip_data = pd.DataFrame(data)
        th_norm = reg_th.predict(ip_data)
        throughoverall = (th_norm * ThroughputMaxmin) + ThroughputMean
        if finalthroughput < throughoverall:
            finalgrpsta = [x for x in grpsta]
            finalgrpdur = [x for x in grpdur]
            finalslotspergrp = [x for x in slotspergrp]
    # running the threads with the updated parameters
    grpsta = [x for x in finalgrpsta]
    grpdur = [x for x in finalgrpdur]
    slotspergrp = [x for x in finalslotspergrp]
    for i in range(len(grpsta)):
        # Stations present in each slot
        slotsta = [[] for j in range(slotspergrp[i])]
        for j in range(len(grpsta[i])):
            slotsta[j % slotspergrp[i]].append(grpsta[i][j])
        # Time duration of each raw slot
        pktslot = []
        for j in range(slotspergrp[i]):
            prec = 0
            for k in slotsta[j]:
                prec += pktrec[k]
            else:
                pktslot.append(prec)
        else:
            invpktslot = [1 if x == 0 else float(1/x) for x in pktslot]
            slottime = [float((grpdur[i] * x)/sum(invpktslot)) for x in invpktslot]
        # Iterating in each RAW slot
        for j in range(slotspergrp[i]):
            # running stations in a slot no. of threads parallelly
            for k in slotsta[j]:
                stationthreads = threading.Thread(target = mac, name = str(i), args=(lock, k, slottime[j],))
                stationthreads.start()
            else:
                # to resume execution of main function after last thread is completed
                stationthreads.join()
    else:
        notimes += 1
        print(str(notimes) + " time")
        print("grpsta: ", end = '')
        print(grpsta)
        print("pktrec ", end = '')
        print(pktrec)
        fairness[notimes - 1] = 1 - (statistics.stdev(pktrec)/statistics.mean(pktrec))
        # sheet.cell(row = notimes, column=1).value = fairness[notimes-1]
        # print("Estimated Throughput: ", end = '')
        # print(finalthroughput)
        print("Observed throughput: ", end = '')
        print((sum(pktrec) * 256 * 8)/(notimes * 0.65 * 1024 * 1024 * beaconinterval))
        print('')
    itr -= 1
    avg = (sum(pktrec) * 256 * 8)/(notimes * 0.65 * 1024 * 1024 * beaconinterval)
    # sheet.cell(row = notimes, column=3).value = avg
else:
    plt.title("Fairness")
    plt.xlabel("Beacon Intervals")
    plt.ylabel("Std deviation (packets)")
    x = [i for i in range(20)]
    plt.plot(x, fairness)
    plt.ylim(0,1)
    plt.show()
    print("average throughput: ", end = '')
    print(avg)
    print('finished')
    # wb.save('fairness_100.xlsx')
