import random
import time
import threading
import statistics
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Reading the excel file
# import pandas as pd
# df = pd.read_excel('dataset.xlsx')
# df_norm = (df - df.mean())/(df.max() - df.min())
# # Mean values of dataset
# stationMean = 265
# slotsMean = 25.642857
# BeaconIntervalMean = 121192.087912
# ThroughputMean = 0.278645
# # max - min of Dataset
# stationMaxmin = 470
# slotsMaxmin = 49
# BeaconIntervalMaxmin = 163840
# ThroughputMaxmin = 0.264033
# # Dataset for no of slots
# slots_x = df_norm[['No of Stations', 'Beacon Interval', 'Throughput']]
# slots_y = df_norm[['No of Slots']]
# # Dataset for Throughput
# th_x = df_norm[['No of Stations', 'Beacon Interval', 'No of Slots']]
# th_y = df_norm[['Throughput']]
# # Regression models
# from sklearn.tree import DecisionTreeRegressor
# reg_slots = DecisionTreeRegressor(max_depth = 7)
# reg_th = DecisionTreeRegressor(max_depth = 20)
# reg_slots.fit(slots_x, slots_y)
# reg_th.fit(th_x, th_y)

# Threading function call for each station
def mac(lock, station_id, slottime):
    begin = time.time()  # time when thread got started
    global pktrec       #  Packets recieved for each station
    global channel      #  Channel is set as global
    CWmin = 15          #  Min window size
    CWmax = 1024        #  Max window size
    while True:
        if time.time() - begin > slottime:
            break
        count = 0         # a variable to keep track of retrylimit and use in selecting backoff value
        while True:
            if time.time() - begin > slottime:
                break
            # Selecting a random backoff counter value
            count += 1
            backOff = int(random.randint(0,min((pow(2,count)-1) * CWmin, CWmax)))
            while backOff > 0:
                if time.time() - begin > slottime:
                    break
                # Waste is for sigma is taken as 50us
                waste = [0 for i in range(1000)]

                if channel == 0:
                    backOff = backOff - 1
            else:
                while channel == 1:
                    if time.time() - begin > slottime:
                        break
                    waste = [0 for i in range(1000)]
                if time.time() - begin > slottime:
                    break
                lock.acquire()
                if channel == 0:
                    channel = 1
                else:
                    while channel:
                        waste = [0 for i in range(1000)]
                    else:
                        channel = 1
                lock.release()
                if time.time() - begin > slottime:
                    channel = 0
                    break
                time.sleep(250 * (10**(-6)))    #  Time for sending the packet
                if time.time() - begin > slottime:
                    channel = 0
                    break
                pktrec[station_id] += 1
                channel = 0

# Number of Stations
N = int(input("No of Nodes connected to the server: "))
# Data Rate
# datarate = float(input("Enter the data rate in Mbps: "))
# data rate is taken as 0.65 Mbps
# Packet Size
size = int(input("Enter the packet size in bytes: "))
# Packet Arrival Rate
# Saturates traffic scenario is considered
# pktrate = int(input("Enter packet arrival rate: "))
# Beacon Interval
# beaconinterval = int(input("Enter Beacon Interval in us: "))
# beaconinterval = float(beaconinterval * (10 ** (-6)))
beaconinterval = 1.0
# Channel
channel = 0
# 0 means channel idle
# 1 means channel busy
pktrec = [0 for i in range(N)]
# Initialisation
# No of groups
# print("For Initialisation")
# nogrp = int(input("No of groups to be present: "))

# wb = load_workbook('fairness_100.xlsx')
# sheet = wb.active
# wb.save('station.xlsx')


fairness = [0 for x in range(20)]
i = 0
lock = threading.Lock()
avg = 0

while i != 20:
    for k in range(N):
        stationthreads = threading.Thread(target = mac, name = str(k), args=(lock, k, beaconinterval,))
        stationthreads.start()
    else:
        # to resume execution of main function after last thread is completed
        stationthreads.join()
    print("pktrec: ", end = '')
    print(pktrec)
    print(i+1)
    fairness[i] = 1 - (statistics.stdev(pktrec)/statistics.mean(pktrec))
    # sheet.cell(row = i+1, column=2).value = fairness[i]
    i += 1
    avg = (sum(pktrec) * 256 * 8)/((i+1) * 0.65 * 1024 * 1024 * beaconinterval)
    print("Throughput: ", end = '')
    print(avg)
    # sheet.cell(row = i+1, column=4).value = avg
    print('')
else:
    plt.title("Fairness")
    plt.xlabel("Beacon Intervals")
    plt.ylabel("Std deviation (packets)")
    x = [i for i in range(20)]
    plt.plot(x, fairness)
    ply.ylim(0,1)
    plt.show()
    print("average throughput: ", end = '')
    print(avg)
    print('finished')
    # wb.save('fairness_100.xlsx')
