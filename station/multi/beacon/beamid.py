import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt

bealis = [0.5 + 0.1*(i) for i in range(11)]
wb = load_workbook('beacon.xlsx')
sheet = wb.active
for i in range(len(bealis)):
    thr.append(float(sheet.cell(row = i+1, column=2).value))
else:
    plt.title("Throughput vs Beacon")
    plt.xlabel("No of stations")
    plt.ylabel("Beacon Interval")
    plt.plot(bealis, thr)
    plt.show()
    print('finished')
