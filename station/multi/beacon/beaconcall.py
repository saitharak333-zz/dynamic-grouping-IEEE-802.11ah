import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt

bea = 0.5
bealis = []
thr = []
count = 1
while bea <= 1.5:
    bealis.append(bea)
    cmd = 'python3 beaconmain.py'
    cmd = cmd + ' ' + str(count) + ' ' + str(bea)
    os.system(cmd)
    bea = bea + 0.1
    count += 1
else:
    wb = load_workbook('beacon.xlsx')
    sheet = wb.active
    for i in range(len(bealis)):
        thr.append(float(sheet.cell(row = i+1, column=2).value))
    else:
        plt.title("Throughput vs Beacon")
        plt.xlabel("No of stations")
        plt.ylabel("Beacon Interval")
        plt.plot(bealis, thr)
        plt.show()
        print('finished')
